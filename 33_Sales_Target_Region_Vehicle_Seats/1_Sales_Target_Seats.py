import os
import xlsxwriter

##########################################################################################################################################################################################################

workbook = xlsxwriter.Workbook('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_result.xlsx')
workbook.close()

##########################################################################################################################################################################################################

import html2text

# Source site 1: # Web site: https://www.ukmarketingmanagement.com/email-lists/uk-manufacturers-list/car-manufacturers-list/
list_brands_1 = open(
    "C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\HTML_Data\\Bucket_Seats_All_Listcarbrands_web.txt")

import csv
import pandas as pd
import xlsxwriter

for line in list_brands_1:
    df = print(html2text.html2text(line))
list_brands_1.close()

print(list_brands_1)

import html2text
import re

html = open("C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\HTML_Data\\Bucket_Seats_All_Listcarbrands_web.txt")
f = html.read()
w = open("C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_output.txt", "wb")
w.write(html2text.html2text(f).encode('utf-8'))
html.close()
w.close()

# Remove special characters from txt files using Python
string = open('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_output.txt').read()
new_str = re.sub('[^a-zA-Z0-9\n\.]', ' ', string)
open('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_Clean_Output.txt', 'w').write(new_str)

# opening and creating new .txt file
with open(
        "C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_Clean_Output.txt", 'r') as r, open(
    'C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_output.txt', 'w') as o:
    for line in r:
        # isspace() function
        if not line.isspace():
            o.write(line)

f = open("C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_output.txt", "r")
print("New text file:\n", f.read())
# Close the files
f.close()

import glob
import openpyxl
def text_into_spreadsheet():
    """main logic for read .txt into spreadsheet"""
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    for column, filename in enumerate(glob.iglob("C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_output.txt"), start=1):
        with open(filename) as textfile:
            sheet.cell(row=1, column=column).value = filename

            for row, line in enumerate(textfile, start=2):
                sheet.cell(row=row, column=column).value = line

    workbook.save('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_result.xlsx')

if __name__ == "__main__":
    text_into_spreadsheet()

# import the openpyxl library
import openpyxl

# open the Excel file
workbook = openpyxl.load_workbook('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_result.xlsx')

# select the sheet to modify
sheet = workbook['Sheet']

# change the header name
sheet.cell(row=1, column=1).value = 'OEMs_List'

# save the changes
workbook.save('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_result.xlsx')

# import pandas lib as pd
import pandas as pd

# read by default 1st sheet of an excel file
df = pd.read_excel('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_result.xlsx')

df.OEMs_List = df.OEMs_List.str.strip()

df.to_excel('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_result.xlsx', index=False)

import pandas as pd

df1 = pd.read_excel('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\Final_All_ONS_List.xlsx')
df2 = pd.read_excel('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\1_result.xlsx')

# Delete specific words from column:
df2['OEMs_List'] = df2['OEMs_List'].str.replace("listcarbrands.com", '')
df2['OEMs_List'] = df2['OEMs_List'].str.replace("logo", '')
df2['OEMs_List'] = df2['OEMs_List'].str.replace("https", '')
df2['OEMs_List'] = df2['OEMs_List'].str.replace("Makes", '')

# Apply uppercase to a column in Pandas dataframe
df1['OEMs_List'] = df1['OEMs_List'].str.upper()
# Apply uppercase to a column in Pandas dataframe
df2['OEMs_List'] = df2['OEMs_List'].str.upper()

# Remove duplicate words in the same cell within a column in python
df2['OEMs_List'] = df2['OEMs_List'].str.split().apply(set).str.join(" ")

# DataFrames
df3 = df1._append(df2,ignore_index=True)

df3 = df3.drop_duplicates()

# Apply uppercase to a column in Pandas dataframe
df3['OEMs_List'] = df3['OEMs_List'].str.upper()
df3 = df3.sort_values("OEMs_List")

#Remove non relevant words:
df3['OEMs_List'] = df3['OEMs_List'].str.replace("Makes that start with the letter F ", '')
df3['OEMs_List'] = df3['OEMs_List'].str.replace("WWW.CARSGUIDE.COM.AU", '')

print(df3)

# exports the dataframe into excel file with
# specified name.
df3.to_excel('C:\\Users\\alvesd\\OneDrive - smmt.co.uk\\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats\\Final_All_ONS_List.xlsx', index=False)

print('Step - Delete the old files from the directory:')
path = r'C:\\Users\\alvesd\\OneDrive - smmt.co.uk\Desktop\\Diego_work_folder\\python\\33_Sales_Target_Region_Vehicle_Seats'
os.chdir(path)
for file in os.listdir(path):
    if file.endswith('.txt'):
        # print(file)
        os.remove(file)

print(html2text.html2text(""""""))
print(html2text.html2text(""""""))
print(html2text.html2text(""""""))

print('Done up to here..............')
#######################################