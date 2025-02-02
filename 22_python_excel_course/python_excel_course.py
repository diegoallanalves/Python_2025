'''
#How to load excel data, add columns headers and save the new data to a folder.
import pandas as pd
from openpyxl.workbook import workbook

df_excel = pd.read_excel('regions.xlsx')
df_csv = pd.read_csv('Names.csv', header=None)
df_txt = pd.read_csv('data.txt', delimiter='\t')

df_csv.columns = ['First', 'Last', 'Address', 'City', 'State', 'Area Code']
df_csv.to_excel('modified.xlsx')
print(df_csv)

#Select specific rows or columns in the data:
import pandas as pd
from openpyxl.workbook import workbook
df = pd.read_csv('Names.csv', header=None)
df.columns = ['First', 'Last', 'Address', 'City', 'State', 'Area Code']
print(df['First'][0:3])

#Print the full description of one single individual:
print(df.iloc[1]) # first add the functions and then the individual location
#Find the location with coordenations:
print(df)

print(df.iloc[5,1]) # first add the functions and then the individual location
#Lets creat a new data by selecting only some columns from the main data:
wanted_values = df[['First', 'Last', 'State']]
extracted_values = wanted_values.to_excel('extracted_values.xlsx', index=None)
print(extracted_values)

###Income data analysis:
import pandas as pd
df_wages = pd.read_csv('Names_wages.csv', header=None)
df_wages.columns = ['First', 'Last', 'Address', 'City', 'State', 'Area Code', 'Income']
print(df_wages)
print(df_wages.loc[df_wages['City'] == 'Riverside'])
#Look or John from Riverside with wages below 4500
print(df_wages.loc[(df_wages['City'] == 'Riverside') & (df_wages['First'] == 'John')])
#Apply tax percentage colunm:
df_wages['Tax %'] = df_wages['Income'].apply(lambda x: .15 if 10000 < x < 40000 else .2 if 40000 < x < 80000 else .25)
print(df_wages)

# See the total ammount of taxes that are paying based on income and %:
df_wages['Taxes Owed'] = df_wages['Income'] * df_wages['Tax %']
print(df_wages['Taxes Owed'])

# Drop some colounms to help visualised the data from a Data Science point:
to_drop = ['Area Code', 'First', 'Address']
df_wages.drop(columns=to_drop, inplace=True)
print(df_wages)

# Find out wha wages bellow 60000:
df_wages['Bellow 60000'] = False
df_wages.loc[df_wages['Income'] < 60000, 'Bellow 60000'] = True
print(df_wages)
# Group by:
print(df_wages.groupby(['Bellow 60000']).mean())
# Sort by income:
print(df_wages.groupby(['Bellow 60000']).mean().sort_values('Income'))

# Clean data:
# Import Libraries
import pandas as pd
import numpy as np
from openpyxl.workbook import Workbook

df = pd.read_csv('Names_wages.csv', header=None)
df.columns = ['First', 'Last', 'Address', 'City', 'State', 'Area Code', 'Income']

# Drop the columns:
df.drop(columns='Address', inplace=True)

# Set location code:
df = df.set_index('Area Code')
# ind information based on the location code:
print(df.loc[8074:, 'First']) # Notice that two Johns appeared in the results. Next step, split the data.
df.First = df.First.str.split(expand=True)
print(df)

# Replace N/A values:
df = df.replace(np.nan, 'N/A', regex=True)
# Now save to excel:
to_excel = df.to_excel('modified.xlsx')

# Learn how to use openpyxl, this creates a new sheet in excel:

from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active

ws1 = wb.create_sheet('NewSheet')
ws2 = wb.create_sheet('Another', 0)

ws.title = 'MtSheet'

wb2 = load_workbook('regions.xlsx')

new_sheet = wb2.create_sheet('NewSheet')
active_sheet = wb2.active

cell = active_sheet['A1']
active_sheet['A1'] = 0
wb2.save('modified.xlsx')

### Work around an excel file:
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook('regions.xlsx')
ws = wb.active

#How to access column range:
cell_range = ws['A1':'C1']
col_range = ws['A':'C']
print(col_range)

#How to access row range, if want to access rows 1 to 5:
row_range = ws[1:5]
print(row_range)

### Loop into columns and rows:
from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook('regions.xlsx')
ws = wb.active
for row in ws.iter_rows(min_row=1, max_col=3, max_row=2, values_only=True):
    for cell in row:
        print(cell)


### How to build graphics and charts using python to excel data:
#First import the excel library:
import openpyxl
from openpyxl.chart import PieChart, Reference, Series, PieChart3D
# Create a data to practice:
wb = openpyxl.Workbook()
ws = wb.active
# Build an icecream shop and show icecream lavour that sells performance:
data = [['Flavor', 'Sold'],
        ['Vanilla', 1500],
        ['Choclate', 1700],
        ['Strawberry', 900],
        ['Banana', 950],
        ]

for rows in data:
    ws.append(rows)

#Now create a chart Variable:
chart = PieChart()
labels = Reference(ws, min_col=1, min_row=2, max_row=5)
data = Reference(ws, min_col=2, min_row=1, max_row=5)
chart.add_data(data, titles_from_data=True)
chart.set_categories(labels)
chart.title = 'Ice Cream by Flavour'

ws.add_chart(chart, 'D1')
wb.save('Pie.xlsx')

### This task gives colours to the excel colunm:
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl import load_workbook

wb = load_workbook('Pie.xlsx')
ws = wb.active

tab = Table(displayName='Table1', ref='A1:B5')
style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False, showLastColumn=False,
                       showRowStripes=True, showColumnStripes=True)
tab.tableStyleInfo = style
ws.add_table(tab)
wb.save('table.xlsx')

# Now lets add image to the worksheet:
img = Image('madecraft.jpg')
img.height = img.height * 1.25 #this code changes the picture height
img.width = img.width * 1.45 #this cde changes the pictures width
ws.add_image(img, 'M1')
wb.save('image.xlsx')

'''
# This code joints excel's documents:
import pandas as pd

df_1 = pd.read_excel('shifts.xlsx', sheet_name='Sheet')
df_2 = pd.read_excel('shifts.xlsx', sheet_name='Sheet1')
df_3 = pd.read_excel('shift_3.xlsx')
df_all = pd.concat([df_1, df_2, df_3], sort=False)
print(df_all)
print(df_all.loc[50]) # This prints all the rows at 50 location.
print(df_all.groupby(['Shift']).mean()['Units Sold']) # Print the mean of "Units sold" per shift.
